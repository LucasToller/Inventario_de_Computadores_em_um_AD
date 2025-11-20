import os
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill

# Gera TXT individual com todos os dados coletados por computador
def gerar_txt_por_computador(dados: dict, pasta_txts: str, computador: str, data_execucao: str) -> str:
        nome_txt = f"{computador}_{data_execucao}.txt"
        caminho_txt = os.path.join(pasta_txts, nome_txt)
        os.makedirs(pasta_txts, exist_ok=True)
        with open(caminho_txt, "w", encoding="utf-8") as txt:
            txt.write(f"{computador}  {data_execucao}\n{'='*60}\n\n")
            
            txt.write("[USUARIO]\n")
            txt.write(f"Usuario Atual Logado: {dados.get('Usuario_Logado')}\n")
            txt.write(f"Nome Completo do Usuario: {dados.get('Usuario_Nome_Completo')}\n\n")
            
            txt.write("[IDENTIFICACAO]\n")
            txt.write(f"Fabricante: {dados.get('Fabricante')}\n")
            txt.write(f"Modelo: {dados.get('Modelo')}\n")
            txt.write(f"Numero de Serie: {dados.get('Numero_Serie')}\n\n")
            
            txt.write("[SISTEMA OPERACIONAL]\n")
            txt.write(f"Sistema Operacional: {dados.get('SO')}\n")
            txt.write(f"Versao do SO: {dados.get('Versao_SO')}\n")
            txt.write(f"Data de Instalacao do SO: {dados.get('Data_Instalacao_SO')}\n\n")
            
            txt.write("[BOOT]\n")
            txt.write(f"Ultimo Boot: {dados.get('Ultimo_Boot')}\n")
            txt.write(f"Tempo de Inicializacao: {dados.get('Tempo_Inicializacao')}\n")
            txt.write(f"Tempo Total Ligado: {dados.get('Tempo_Total_Ligado')}\n\n")
            
            txt.write("[PROCESSADOR]\n")
            txt.write(f"Processador: {dados.get('Processador')}\n\n")
            
            txt.write("[PLACA DE VIDEO]\n")
            txt.write(f"Placa de Video: {dados.get('Placa_Video')}\n\n")
            
            txt.write("[MEMORIA RAM]\n") 
            modulos_RAM = dados.get("Memoria_RAM", [])
            if modulos_RAM:
                for i, m in enumerate(modulos_RAM, start=1):
                    capacidade_RAM = m.get("Capacidade_GB", "N/D")
                    frequencia_RAM = m.get("Frequencia_MHz", "N/D")
                    geracao_RAM = m.get("Geracao_RAM", "N/D")
                    txt.write(f"Modulo {i}: {capacidade_RAM} GB - {frequencia_RAM} MHz - {geracao_RAM}\n")
                txt.write(f"Quantidade de modulos: {dados.get('Qtd_Modulos_RAM')}\n")
                txt.write(f"Geracao da memoria: {dados.get('Geracao_RAM')}\n")
                txt.write(f"Memoria total: {dados.get('Total_GB_RAM')} GB\n\n")
            else:
                txt.write("Nenhum modulo de memoria detectado.\n\n")
            
            txt.write("[ARMAZENAMENTO]\n")
            txt.write(f"Armazenamento Total do Disco C: {dados.get('DiscoC_Total_GB')} GB\n")
            txt.write(f"Utilizado: {dados.get('DiscoC_Usado_GB')} GB ({dados.get('DiscoC_Uso_Pct')}%)\n")
            txt.write(f"Disponivel: {dados.get('DiscoC_Livre_GB')} GB ({dados.get('DiscoC_Livre_Pct')}%)\n")
            txt.write(f"Tipo de Armazenamento: {dados.get('Tipo_Armazenamento')}\n\n")
            
            txt.write("[REDE]\n")
            txt.write(f"IP Principal: {dados.get('IP_Principal')}\n")
            txt.write(f"MAC: {dados.get('MAC')}\n")
            txt.write(f"Velocidade de Rede: {dados.get('Velocidade_Rede')}\n")
            txt.write(f"Placa de Rede: {dados.get('Placa_Rede')}\n\n")
                
        # Fim da funcao gerar_txt_por_computador        
        return caminho_txt


# Gera planilha - Somente apos finalizar a coleta dos dados de todos os computadores
def gerar_planilha(todos_dados: list[dict], caminho_xlsx: str) -> None:
    planilha = Workbook()
    aba_planilha = planilha.active
    aba_planilha.title = "Inventario"

    campos_planilha = [
        "Usuario_Logado", "Usuario_Nome_Completo",
        "Computador", "Fabricante", "Modelo", "Numero_Serie",
        "SO", "Versao_SO", "Data_Instalacao_SO", 
        "Ultimo_Boot", "Tempo_Inicializacao", "Tempo_Total_Ligado",
        "Processador", "Placa_Video",
        "Total_GB_RAM", "Qtd_Modulos_RAM", "Geracao_RAM",
        "DiscoC_Total_GB","DiscoC_Usado_GB","DiscoC_Uso_Pct","DiscoC_Livre_GB","DiscoC_Livre_Pct","Tipo_Armazenamento",
        "IP_Principal","MAC","Velocidade_Rede","Placa_Rede",
    ]
    aba_planilha.append(campos_planilha)

    # Estilo do cabecalho
    for celula in aba_planilha[1]:
        celula.font = Font(bold=True)
        celula.fill = PatternFill("solid", fgColor="DDDDDD")
        celula.alignment = Alignment(horizontal="center", vertical="center")

    # Mapas de formatacao por coluna
    colunas_duas_casas  = {"Total_GB_RAM", "DiscoC_Total_GB", "DiscoC_Usado_GB", "DiscoC_Livre_GB"}
    colunas_inteiro     = {"Qtd_Modulos_RAM"}
    colunas_percentual  = {"DiscoC_Uso_Pct", "DiscoC_Livre_Pct"}

    indice_por_campo = {nome: i for i, nome in enumerate(campos_planilha, start=1)}
    largura_maxima_coluna = {i: len(nome) for nome, i in indice_por_campo.items()}  

    # Congela o cabecalho da planilha
    aba_planilha.freeze_panes = "A2"

    # Escrita das linhas: formata texto, numero, percentual e mede largura
    for registro in todos_dados:
        linha_atual = aba_planilha.max_row + 1
        for nome_campo, indice_coluna in indice_por_campo.items():
            valor_bruto = registro.get(nome_campo, "")
            valor_final = valor_bruto
            numerico = False

            # Conversao para numero quando necessario
            if valor_final == "N/D":
                pass
            elif isinstance(valor_final, (int, float)):
                numerico = True
            elif isinstance(valor_final, str):
                
                # Sanitiza strings que poderiam virar formula no Excel
                if valor_final and valor_final[0] in ("=", "+", "-", "@"):
                    valor_final = "'" + valor_final
                else:
                    # Converter string numerica (suporta virgula como decimal)
                    texto_para_conversao = valor_final.replace(",", ".").strip()
                    try:
                        valor_final = float(texto_para_conversao)
                        numerico = True
                    except ValueError:
                        pass

            celula = aba_planilha.cell(row = linha_atual, column = indice_coluna, value = valor_final)

            # Formatacao numerica por coluna
            if nome_campo in colunas_percentual and numerico:
                
                # Converte para fracao
                if celula.value > 1:
                    celula.value = celula.value / 100.0
                celula.number_format = "0.00%"
            elif nome_campo in colunas_duas_casas and numerico:
                celula.number_format = "0.00"
            elif nome_campo in colunas_inteiro and numerico:
                celula.number_format = "0"

            # Mede largura da coluna 
            texto_celula = str(celula.value) if celula.value is not None else ""
            if len(texto_celula) > largura_maxima_coluna[indice_coluna]:
                largura_maxima_coluna[indice_coluna] = len(texto_celula)
    
    # Habilita AutoFiltro na planilha apos inserir os dados coletados
    aba_planilha.auto_filter.ref = aba_planilha.dimensions

    # Aplica larguras
    for indice_coluna, largura in largura_maxima_coluna.items():
        letra_coluna = get_column_letter(indice_coluna)
        aba_planilha.column_dimensions[letra_coluna].width = min(largura + 2, 60)

    # Salva arquivo
    os.makedirs(os.path.dirname(caminho_xlsx) or ".", exist_ok=True)
    planilha.save(caminho_xlsx)